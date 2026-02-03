import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ACCOUNTING_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df["Service"] = df["Service"].str.replace(r"\s*\d+$", "", regex=True)
    df["Unit"] = df["Unit"].fillna("")
    df["Specifications"] = df["Specifications"].fillna("")
    df = df.groupby(
        ["Service", "Region", "AZ", "Billing Mode", "Purchase Amount", "Unit", "Specifications"]
    )[["Quantity", "Discounted Price (USD)"]].sum().reset_index()

    priority_list = ["Elastic Cloud Server", "Flexus X Instance", "Elastic Volume Service"]
    df["priority_order"] = df["Service"].map(
        lambda value: priority_list.index(value) if value in priority_list else 999
    )
    df = df.sort_values(by=["priority_order", "Service"]).drop(columns="priority_order")

    df["Monthly Price"] = df.apply(
        lambda row: -1 if "RI" in row["Billing Mode"] else row["Discounted Price (USD)"],
        axis=1,
    )
    df["Yearly Price"] = df.apply(
        lambda row: row["Discounted Price (USD)"] if "RI" in row["Billing Mode"] else -1,
        axis=1,
    )
    df = df.drop(columns=["AZ", "Billing Mode", "Purchase Amount", "Unit", "Discounted Price (USD)"])
    df["Comments"] = ""
    return df


def write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, title: str) -> int:
    safe_name = sheet_name[:31]
    df.insert(0, "No", range(1, len(df) + 1))
    n_rows = len(df)

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    df.to_excel(writer, index=False, sheet_name=safe_name, startrow=1)
    ws = writer.book[safe_name]
    ws.sheet_view.showGridLines = False

    n_cols = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, name="Arial", size=20)
    title_cell.alignment = align
    ws.row_dimensions[1].height = 36
    for col in range(1, n_cols + 1):
        ws.cell(row=1, column=col).border = thin

    header_row = 2
    data_start = header_row + 1
    data_end = data_start + n_rows - 1
    total_row = data_end + 1
    for row in range(header_row, total_row + 2):
        ws.row_dimensions[row].height = 35

    headers = [cell.value for cell in ws[header_row]]
    idx_month = headers.index("Monthly Price") + 1
    idx_year = headers.index("Yearly Price") + 1
    idx_comm = headers.index("Comments") + 1

    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=idx_month - 1)
    label_cell = ws.cell(row=total_row, column=1)
    label_cell.value = "Total"
    label_cell.font = Font(bold=True, name="Arial", size=14)
    label_cell.alignment = align
    for col in range(1, idx_month):
        ws.cell(row=total_row, column=col).border = thin

    ws.merge_cells(
        start_row=total_row + 1, start_column=1, end_row=total_row + 1, end_column=idx_month - 1
    )
    for col in range(1, idx_month):
        ws.cell(row=total_row + 1, column=col).border = thin

    col_month = get_column_letter(idx_month)
    month_cell = ws.cell(row=total_row, column=idx_month)
    month_cell.value = f"=SUM({col_month}{data_start}:{col_month}{data_end})"
    month_cell.font = Font(bold=True, name="Arial", size=14)
    month_cell.alignment = align
    month_cell.number_format = ACCOUNTING_USD
    month_cell.border = thin

    col_year = get_column_letter(idx_year)
    year_cell = ws.cell(row=total_row, column=idx_year)
    year_cell.value = f"=SUM({col_year}{data_start}:{col_year}{data_end})"
    year_cell.font = Font(bold=True, name="Arial", size=14)
    year_cell.alignment = align
    year_cell.number_format = ACCOUNTING_USD
    year_cell.border = thin

    ws.cell(row=total_row, column=idx_comm).border = thin

    for row in ws.iter_rows(min_row=data_start, max_row=data_end):
        for cell in row:
            if cell.column == idx_month and cell.value == -1:
                ref = f"{get_column_letter(idx_year)}{cell.row}"
                cell.value = f"=ROUND({ref}/12, 2)"
            elif cell.column == idx_year and cell.value == -1:
                ref = f"{get_column_letter(idx_month)}{cell.row}"
                cell.value = f"=ROUND({ref}*12, 2)"
            cell.font = Font(name="Arial", size=14)
            cell.alignment = align
            cell.border = thin
            if cell.column in (idx_month, idx_year):
                cell.number_format = ACCOUNTING_USD

    for col_idx, _ in enumerate(headers, 1):
        max_len = max(
            len(str(ws.cell(row, column=col_idx).value or ""))
            for row in range(1, total_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 73)

    fill = PatternFill(start_color="003366", fill_type="solid")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=14)
        cell.fill = fill
        cell.alignment = align
        cell.border = thin
    return n_rows


def create_workbook(file_streams, titles, output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pages = []
        for stream, title in zip(file_streams, titles):
            df_raw = pd.read_excel(stream, sheet_name=1, skiprows=1, skipfooter=3)
            df_proc = process_dataframe(df_raw)
            safe = title[:31]
            count = write_sheet(writer, df_proc, sheet_name=safe, title=title)
            pages.append((safe, count))

        wb = writer.book
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        align = Alignment(wrap_text=True, vertical="center", horizontal="center")
        summary = wb.create_sheet(title="Summary")
        writer.sheets["Summary"] = summary
        summary.sheet_view.showGridLines = False

        n_cols = 5
        summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = summary.cell(row=1, column=1)
        title_cell.value = "- Huawei Cloud"
        title_cell.font = Font(bold=True, name="Calibri", size=20)
        title_cell.alignment = align
        summary.row_dimensions[1].height = 36
        for col in range(1, n_cols + 1):
            summary.cell(row=1, column=col).border = thin

        headers = ["No", "Name", "Quotation Link", "Yearly Price (US$)", "Monthly Price (US$)"]
        summary.row_dimensions[2].height = 35
        for idx, header in enumerate(headers, 1):
            cell = summary.cell(row=2, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
            cell.fill = PatternFill(start_color="003366", fill_type="solid")
            cell.alignment = align
            cell.border = thin

        for i, (sheet, count) in enumerate(pages, start=1):
            row = i + 2
            summary.row_dimensions[row].height = 35

            summary.cell(row=row, column=1, value=i).font = Font(name="Calibri", size=12)
            summary.cell(row=row, column=1).alignment = align
            summary.cell(row=row, column=1).border = thin

            summary.cell(row=row, column=2, value=sheet).font = Font(
                name="Calibri", size=12, bold=True
            )
            summary.cell(row=row, column=2).alignment = align
            summary.cell(row=row, column=2).border = thin

            link = summary.cell(row=row, column=3)
            link.value = "Link"
            link.hyperlink = f"#{sheet}!A{count + 4}"
            link.font = Font(
                underline="single", color="0000FF", name="Calibri", bold=True, size=12
            )
            link.alignment = align
            link.border = thin

            ws = wb[sheet]
            header_values = [cell.value for cell in ws[2]]
            idx_year = header_values.index("Yearly Price") + 1
            idx_month = header_values.index("Monthly Price") + 1
            total_row = count + 3

            year_cell = summary.cell(row=row, column=4)
            year_cell.value = f"='{sheet}'!{get_column_letter(idx_year)}{total_row}"
            year_cell.number_format = ACCOUNTING_USD
            year_cell.font = Font(name="Calibri", bold=True, size=12)
            year_cell.alignment = align
            year_cell.border = thin

            month_cell = summary.cell(row=row, column=5)
            month_cell.value = f"='{sheet}'!{get_column_letter(idx_month)}{total_row}"
            month_cell.number_format = ACCOUNTING_USD
            month_cell.font = Font(name="Calibri", bold=True, size=12)
            month_cell.alignment = align
            month_cell.border = thin

        for col_idx in range(1, n_cols + 1):
            max_len = max(
                len(str(summary.cell(row, col_idx).value or "")) for row in range(1, len(pages) + 3)
            )
            summary.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 73)


def main() -> int:
    if len(sys.argv) < 2:
        print("Missing payload path.", file=sys.stderr)
        return 1

    payload_path = Path(sys.argv[1])
    data = json.loads(payload_path.read_text(encoding="utf-8"))
    files = data.get("files", [])
    output_path = Path(data.get("output", "combined-excel.xlsx"))

    if not files:
        print("No files provided.", file=sys.stderr)
        return 1

    streams = [item["path"] for item in files]
    titles = [item.get("title") or Path(item["path"]).stem for item in files]

    create_workbook(streams, titles, output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
