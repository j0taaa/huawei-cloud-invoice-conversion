import io
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ACCOUNTING_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

def process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df["Service"] = df["Service"].str.replace(r"\s*\d+$", "", regex=True)
    df["Unit"] = df["Unit"].fillna("")
    df["Specifications"] = df["Specifications"].fillna("")
    df = df.groupby(
        ["Service", "Region", "AZ", "Billing Mode", "Purchase Amount", "Unit", "Specifications"]
    )[["Quantity", "Discounted Price (USD)"]].sum().reset_index()

    priorityList = ["Elastic Cloud Server", "Flexus X Instance", "Elastic Volume Service"]
    df["priority_order"] = df["Service"].map(lambda x: priorityList.index(x) if x in priorityList else 999)
    df = df.sort_values(by=["priority_order", "Service"]).drop(columns="priority_order")

    df["Monthly Price"] = df.apply(
        lambda x: -1 if "RI" in x["Billing Mode"] else x["Discounted Price (USD)"], axis=1
    )
    df["Yearly Price"] = df.apply(
        lambda x: x["Discounted Price (USD)"] if "RI" in x["Billing Mode"] else -1, axis=1
    )
    df = df.drop(columns=["AZ", "Billing Mode", "Purchase Amount", "Unit", "Discounted Price (USD)"])
    df["Comments"] = ""
    return df

def write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, title: str):
    safe_name = sheet_name[:31]
    df.insert(0, "No", range(1, len(df) + 1))
    n_rows = len(df)

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Dump data
    df.to_excel(writer, index=False, sheet_name=safe_name, startrow=1)
    ws = writer.book[safe_name]
    ws.sheet_view.showGridLines = False

    # Title row
    n_cols = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, name='Arial', size=20)
    title_cell.alignment = align
    ws.row_dimensions[1].height = 36
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).border = thin

    # Row heights
    header_row = 2
    data_start = header_row + 1
    data_end = data_start + n_rows - 1
    total_row = data_end + 1
    for r in range(header_row, total_row + 2):
        ws.row_dimensions[r].height = 35

    # Locate columns
    headers = [c.value for c in ws[header_row]]
    idx_month = headers.index("Monthly Price") + 1
    idx_year  = headers.index("Yearly Price") + 1
    idx_comm  = headers.index("Comments") + 1

    # “Total” label
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=idx_month-1)
    lbl = ws.cell(row=total_row, column=1)
    lbl.value = "Total"
    lbl.font = Font(bold=True, name='Arial', size=14)
    lbl.alignment = align
    for c in range(1, idx_month):
        ws.cell(row=total_row, column=c).border = thin

    ws.merge_cells(start_row=total_row+1, start_column=1, end_row=total_row+1, end_column=idx_month-1)
    for c in range(1, idx_month):
        ws.cell(row=total_row+1, column=c).border = thin

    # Total formulas
    col_m = get_column_letter(idx_month)
    mcell = ws.cell(row=total_row, column=idx_month)
    mcell.value = f"=SUM({col_m}{data_start}:{col_m}{data_end})"
    mcell.font = Font(bold=True, name='Arial', size=14)
    mcell.alignment = align
    mcell.number_format = ACCOUNTING_USD
    mcell.border = thin

    col_y = get_column_letter(idx_year)
    ycell = ws.cell(row=total_row, column=idx_year)
    ycell.value = f"=SUM({col_y}{data_start}:{col_y}{data_end})"
    ycell.font = Font(bold=True, name='Arial', size=14)
    ycell.alignment = align
    ycell.number_format = ACCOUNTING_USD
    ycell.border = thin

    ws.cell(row=total_row, column=idx_comm).border = thin

    # Populate and style data cells
    for row in ws.iter_rows(min_row=data_start, max_row=data_end):
        for cell in row:
            if cell.column == idx_month and cell.value == -1:
                ref = f"{get_column_letter(idx_year)}{cell.row}"
                cell.value = f"=ROUND({ref}/12, 2)"
            elif cell.column == idx_year and cell.value == -1:
                ref = f"{get_column_letter(idx_month)}{cell.row}"
                cell.value = f"=ROUND({ref}*12, 2)"
            cell.font = Font(name='Arial', size=14)
            cell.alignment = align
            cell.border = thin
            if cell.column in (idx_month, idx_year):
                cell.number_format = ACCOUNTING_USD

    # Column widths
    for i, _ in enumerate(headers, 1):
        max_len = max(
            len(str(ws.cell(r, column=i).value or "")) 
            for r in range(1, total_row+1)
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 73)

    # Header styling
    fill = PatternFill(start_color="003366", fill_type="solid")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=14)
        cell.fill = fill
        cell.alignment = align
        cell.border = thin

def create_workbook(file_streams, titles):
    """
    Given a list of file-like objects (each an .xlsx) and their display titles,
    returns a BytesIO containing the combined workbook with data sheets and a Summary.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pages = []
        # Data sheets
        for f, title in zip(file_streams, titles):
            df_raw  = pd.read_excel(f, sheet_name=1, skiprows=1, skipfooter=3)
            df_proc = process_dataframe(df_raw)
            safe    = title[:31]
            write_sheet(writer, df_proc, sheet_name=safe, title=title)
            pages.append((safe, len(df_proc)))

        # Summary sheet
        wb      = writer.book
        thin    = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        align   = Alignment(wrap_text=True, vertical="center", horizontal="center")
        summary = wb.create_sheet(title="Summary")
        writer.sheets["Summary"] = summary
        summary.sheet_view.showGridLines = False

        # Title row
        n_cols = 5
        summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tcell = summary.cell(row=1, column=1)
        tcell.value = "- Huawei Cloud"
        tcell.font  = Font(bold=True, name='Calibri', size=20)
        tcell.alignment = align
        summary.row_dimensions[1].height = 36
        for c in range(1, n_cols+1):
            summary.cell(row=1, column=c).border = thin

        # Header row
        headers = ["No", "Name", "Quotation Link", "Yearly Price (US$)", "Monthly Price (US$)"]
        summary.row_dimensions[2].height = 35
        for idx, h in enumerate(headers, 1):
            cell = summary.cell(row=2, column=idx)
            cell.value = h
            cell.font  = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
            cell.fill  = PatternFill(start_color="003366", fill_type="solid")
            cell.alignment = align
            cell.border = thin

        # Data rows
        for i, (sheet, count) in enumerate(pages, start=1):
            row = i + 2
            summary.row_dimensions[row].height = 35

            # No
            summary.cell(row=row, column=1, value=i).font = Font(name='Calibri', size=12)
            summary.cell(row=row, column=1).alignment = align
            summary.cell(row=row, column=1).border    = thin

            # Name
            summary.cell(row=row, column=2, value=sheet).font = Font(name='Calibri', size=12, bold=True)
            summary.cell(row=row, column=2).alignment = align
            summary.cell(row=row, column=2).border    = thin

            # Link
            link = summary.cell(row=row, column=3)
            link.value     = "Link"
            link.hyperlink = f"#{sheet}!A{count+4}"
            link.font      = Font(underline="single", color="0000FF", name="Calibri", bold=True, size=12)
            link.alignment = align
            link.border    = thin

            # Totals
            ws   = wb[sheet]
            hdrs = [c.value for c in ws[2]]
            idx_y = hdrs.index("Yearly Price") + 1
            idx_m = hdrs.index("Monthly Price") + 1
            total_row = count + 3

            ycell = summary.cell(row=row, column=4)
            ycell.value         = f"='{sheet}'!{get_column_letter(idx_y)}{total_row}"
            ycell.number_format = ACCOUNTING_USD
            ycell.font          = Font(name='Calibri', bold=True, size=12)
            ycell.alignment     = align
            ycell.border        = thin

            mcell = summary.cell(row=row, column=5)
            mcell.value         = f"='{sheet}'!{get_column_letter(idx_m)}{total_row}"
            mcell.number_format = ACCOUNTING_USD
            mcell.font          = Font(name='Calibri', bold=True, size=12)
            mcell.alignment     = align
            mcell.border        = thin

        # Auto-width on Summary
        for col_idx in range(1, n_cols+1):
            max_len = max(
                len(str(summary.cell(r, col_idx).value or "")) 
                for r in range(1, len(pages)+3)
            )
            summary.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 73)

    output.seek(0)
    return output
